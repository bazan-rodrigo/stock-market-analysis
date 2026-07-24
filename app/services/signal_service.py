"""
Servicio de señales.
Evalúa cada SignalDefinition contra indicadores (ind_*) / group_scores
y persiste los resultados en sig_{id} (una tabla por señal, ver
app.models.signal_store) / group_signal_value.
"""
import logging
import sqlalchemy as sa
from datetime import date as date_type

from app.database import get_session
from app.models import (
    Asset,
    GroupScore,
    GroupSignalValue,
    SignalDefinition,
)
from app.models import signal_store
from app.models.indicator_definition import IndicatorDefinition
from app.models.indicator_store import query_values_asof
from app.models.price import Price
from app.services import db_compat, signal_engine

logger = logging.getLogger(__name__)


_VALID_GROUP_INDICATOR_KEYS = frozenset({"regime_score_d", "regime_score_w", "regime_score_m"})

# Indicadores virtuales: no tienen tabla ind_*, se leen de otra fuente
_VIRTUAL_CODES = frozenset({"last_close"})


def _load_virtual(s, code: str, target_date) -> dict:
    """Carga un indicador virtual. Retorna {asset_id: value}."""
    if code == "last_close":
        rows = s.query(Price.asset_id, Price.close).filter(Price.date == target_date).all()
        return {r[0]: float(r[1]) for r in rows if r[1] is not None}
    return {}


def _get_group_indicator_value(gscore: GroupScore, key: str):
    if key not in _VALID_GROUP_INDICATOR_KEYS:
        logger.warning("signal_service: indicator_key '%s' no es un campo válido de GroupScore", key)
        return None
    return getattr(gscore, key)


def _prepare_signals(s, only_signal_ids: set[int] | None = None) -> dict | None:
    """Contexto invariante de una corrida: señales parseadas y clasificadas
    + códigos de indicador que necesitan. Compartido por el camino por-fecha
    (que lo rearma en cada llamada) y el modo rango (una vez por corrida).
    None si no hay señales que evaluar."""
    import json as _json
    from types import SimpleNamespace

    signals_orm = s.query(SignalDefinition).all()
    if only_signal_ids is not None:
        signals_orm = [sg for sg in signals_orm if sg.id in only_signal_ids]
    if not signals_orm:
        return None

    # Copias planas: los evaluadores acceden a estos atributos millones de
    # veces por corrida y el descriptor instrumentado del ORM (~230ns por
    # acceso) pesa de verdad a esa escala
    signals = [
        SimpleNamespace(
            id=sg.id, key=sg.key, source=sg.source, group_type=sg.group_type,
            indicator_key=sg.indicator_key, formula_type=sg.formula_type,
            params=sg.params)
        for sg in signals_orm
    ]

    # Params parseados una sola vez por señal (evita json.loads por activo)
    params_by_id: dict[int, dict | None] = {}
    for sig in signals:
        try:
            params_by_id[sig.id] = _json.loads(sig.params)
        except (TypeError, ValueError):
            params_by_id[sig.id] = None

    asset_signals  = [sg for sg in signals if sg.source == "asset"]
    group_signals  = [sg for sg in signals if sg.source == "group"]

    # Señales de grupo mal configuradas (indicator_key que no es un campo de
    # group_scores): descartarlas UNA vez acá — evaluarlas warnearía por cada
    # (grupo × fecha), inundando el log en un backfill
    bad_group = [sg for sg in group_signals
                 if sg.indicator_key not in _VALID_GROUP_INDICATOR_KEYS]
    if bad_group:
        logger.warning(
            "signal_service: señales de grupo ignoradas por indicator_key "
            "inválido (los scores de grupo solo tienen %s): %s",
            sorted(_VALID_GROUP_INDICATOR_KEYS),
            sorted(sg.key for sg in bad_group))
        group_signals = [sg for sg in group_signals if sg not in bad_group]

    # Descubrir qué indicator_keys necesitan las señales de activo
    needed_codes = {sg.indicator_key for sg in asset_signals if sg.indicator_key}

    # keep_history por código, para decidir de dónde leer cada uno
    keep_history_by_code = {
        d.code: d.keep_history for d in s.query(IndicatorDefinition).all()
    }

    # Evaluadores COMPILADOS una vez por corrida (closures con los params
    # horneados): ver signal_engine.compile_evaluator — el despacho por
    # llamada de evaluate() era ~75% del cómputo puro del backfill.
    compiled_by_id = {
        sig.id: signal_engine.compile_evaluator(
            sig.formula_type, params_by_id.get(sig.id), sig.params)
        for sig in signals
    }

    return {
        "signals":        signals,
        "params_by_id":   params_by_id,
        "compiled_by_id": compiled_by_id,
        "asset_signals":  asset_signals,
        "group_signals":  group_signals,
        "hist_codes":     {c for c in needed_codes - _VIRTUAL_CODES
                           if keep_history_by_code.get(c)},
        "nohist_codes":   {c for c in needed_codes - _VIRTUAL_CODES
                           if keep_history_by_code.get(c) is False},
        "virtual_codes":  needed_codes & _VIRTUAL_CODES,
    }


def compute_signal_values(target_date: date_type,
                          only_signal_ids: set[int] | None = None,
                          latest_price_date: date_type | None = None) -> int:
    """
    Calcula los scores de señal de target_date para todos los activos y
    los upsertea en las tablas sig_{id} (una por señal).
    Lee valores desde cada tabla ind_{code} por separado.

    only_signal_ids acota el cálculo a un subconjunto (alcance por señal o
    estrategia del backfill).

    latest_price_date evita re-consultar MAX(prices.date) (caro sin índice
    por fecha) cuando el llamador itera muchas fechas (backfill).
    """
    s = get_session()

    prep = _prepare_signals(s, only_signal_ids)
    if prep is None:
        return 0
    signals        = prep["signals"]
    params_by_id   = prep["params_by_id"]
    asset_signals  = prep["asset_signals"]
    group_signals  = prep["group_signals"]
    hist_codes     = prep["hist_codes"]
    nohist_codes   = prep["nohist_codes"]
    virtual_to_load = prep["virtual_codes"]

    # Construir {asset_id: {code: value}} leyendo cada ind_* table con
    # lookup as-of (última fila <= target_date): los indicadores
    # semanales/mensuales se guardan con fechas de fin de período, un match
    # exacto los dejaba en 0 scores casi cualquier día (tendencia_w/m,
    # volatilidad_w/m nunca puntuaban)
    isnaps: dict[int, dict] = {}
    for code in hist_codes:
        try:
            values_by_asset = query_values_asof(s, code, target_date)
        except Exception:
            continue
        for asset_id_row, value in values_by_asset.items():
            isnaps.setdefault(asset_id_row, {})[code] = value

    # Indicadores sin historia (drawdown_current, etc.): solo existe el valor
    # vigente en current_indicator_values. Usarlo únicamente cuando
    # target_date ES la fecha vigente — para fechas pasadas sería sesgo de
    # anticipación silencioso, mejor que la señal no puntúe.
    if nohist_codes:
        if latest_price_date is None:
            from app.services.group_score_service import get_default_target_date
            latest_price_date = get_default_target_date()
        if target_date == latest_price_date:
            from app.models.indicator_store import CurrentIndicatorValue
            rows = s.query(
                CurrentIndicatorValue.asset_id, CurrentIndicatorValue.code,
                CurrentIndicatorValue.value_num, CurrentIndicatorValue.value_str,
            ).filter(CurrentIndicatorValue.code.in_(nohist_codes)).all()
            for asset_id_row, code, num, txt in rows:
                value = num if num is not None else txt
                if value is not None:
                    isnaps.setdefault(asset_id_row, {})[code] = value
        else:
            logger.info(
                "signal_service: señales sobre indicadores sin historia (%s) "
                "omitidas para fecha pasada %s (solo evaluables en la fecha "
                "vigente)", sorted(nohist_codes), target_date)

    # Indicadores virtuales (last_close → prices table)
    for code in virtual_to_load:
        for asset_id, value in _load_virtual(s, code, target_date).items():
            isnaps.setdefault(asset_id, {})[code] = value

    if not isnaps:
        logger.info("signal_service: sin valores de indicadores para %s", target_date)
        return 0

    # Cargar group_scores del día
    gscores: dict[tuple, GroupScore] = {
        (gs.group_type, gs.group_id): gs
        for gs in s.query(GroupScore).filter(GroupScore.date == target_date).all()
    }

    # Info de grupo de cada activo
    asset_groups: dict[int, dict] = {
        a.id: {
            "sector":          a.sector_id,
            "market":          a.market_id,
            "industry":        a.industry_id,
            "country":         a.country_id,
            "instrument_type": a.instrument_type_id,
        }
        for a in s.query(
            Asset.id, Asset.sector_id, Asset.market_id,
            Asset.industry_id, Asset.country_id, Asset.instrument_type_id,
        ).all()
    }

    scores = _evaluate_asset_signal_scores(
        compiled_by_id=prep["compiled_by_id"],
        signals=signals, asset_signals=asset_signals,
        group_signals=group_signals, params_by_id=params_by_id,
        isnaps=isnaps, asset_groups=asset_groups, gscores=gscores)

    # Cada señal escribe en su propia tabla sig_{id} (upsert de la fecha)
    by_sig: dict[int, dict[int, float]] = {}
    for (sig_id, asset_id), score in scores.items():
        by_sig.setdefault(sig_id, {})[asset_id] = score

    written = 0
    for sig_id, asset_scores in by_sig.items():
        t = signal_store.ensure_sig_table(sig_id, bind=s.connection())
        existing = {
            aid: sc for aid, sc in s.execute(
                sa.select(t.c.asset_id, t.c.score).where(t.c.date == target_date))
        }
        ins = [{"asset_id": aid, "date": target_date, "score": sc}
               for aid, sc in asset_scores.items() if aid not in existing]
        upd = [{"aid": aid, "sc": sc} for aid, sc in asset_scores.items()
               if aid in existing and existing[aid] != sc]
        if ins:
            s.execute(t.insert(), ins)
        if upd:
            s.execute(
                t.update().where(t.c.date == target_date,
                                 t.c.asset_id == sa.bindparam("aid"))
                .values(score=sa.bindparam("sc")),
                upd)
        written += len(asset_scores)

    s.commit()
    logger.info("signal_service: %d signal_value escritos para %s", written, target_date)
    return written


def _evaluate_asset_signal_scores(*, signals, asset_signals, group_signals,
                                  params_by_id, isnaps,
                                  asset_groups, gscores,
                                  compiled_by_id=None) -> dict[tuple, float]:
    """{(signal_id, asset_id): score} de una fecha — LÓGICA PURA, sin BD,
    compartida por el camino por-fecha y el modo rango (la paridad entre
    ambos depende de que este sea el único evaluador).

    gscores: {(group_type, group_id): obj} con atributos regime_score_*
    (ORM GroupScore o SimpleNamespace, indistinto).

    compiled_by_id: evaluadores compilados de _prepare_signals; si no viene
    (llamadores viejos/tests) se compila acá — mismo resultado, solo se
    paga la compilación (barata) en cada llamada."""
    if compiled_by_id is None:
        compiled_by_id = {
            sig.id: signal_engine.compile_evaluator(
                sig.formula_type, params_by_id.get(sig.id), sig.params)
            for sig in signals
        }
    results: dict[tuple, float] = {}

    # Memo de scores de grupo: todos los activos de un mismo grupo comparten
    # el mismo score, no hace falta evaluarlo una vez por activo.
    group_score_memo: dict[tuple, float | None] = {}
    id_by_key = {sig.key: sig.id for sig in signals}

    for asset_id, isnap in isnaps.items():
        asset_scores: dict[str, float | None] = {}

        for sig in asset_signals:
            value = isnap.get(sig.indicator_key) if sig.indicator_key else None
            asset_scores[sig.key] = compiled_by_id[sig.id](value)

        groups = asset_groups.get(asset_id, {})
        for sig in group_signals:
            group_id = groups.get(sig.group_type)
            if group_id is None:
                asset_scores[sig.key] = None
                continue
            memo_key = (sig.id, group_id)
            if memo_key in group_score_memo:
                asset_scores[sig.key] = group_score_memo[memo_key]
                continue
            gscore = gscores.get((sig.group_type, group_id))
            if gscore is None:
                group_score_memo[memo_key] = None
                asset_scores[sig.key] = None
                continue
            value = _get_group_indicator_value(gscore, sig.indicator_key) if sig.indicator_key else None
            score = compiled_by_id[sig.id](value)
            group_score_memo[memo_key] = score
            asset_scores[sig.key] = score

        for key, score in asset_scores.items():
            if score is None:
                continue
            sig_id = id_by_key.get(key)
            if sig_id is not None:
                results[(sig_id, asset_id)] = score

    return results


def compute_group_signal_values(target_date: date_type,
                                only_signal_ids: set[int] | None = None) -> int:
    s = get_session()

    group_signals = (
        s.query(SignalDefinition).filter(SignalDefinition.source == "group").all()
    )
    if only_signal_ids is not None:
        group_signals = [sg for sg in group_signals if sg.id in only_signal_ids]
    # Mal configuradas (ver compute_signal_values): un solo warning, no por grupo
    bad_group = [sg for sg in group_signals
                 if sg.indicator_key not in _VALID_GROUP_INDICATOR_KEYS]
    if bad_group:
        logger.warning(
            "signal_service: señales de grupo ignoradas por indicator_key "
            "inválido: %s", sorted(sg.key for sg in bad_group))
        group_signals = [sg for sg in group_signals if sg not in bad_group]
    if not group_signals:
        return 0

    import json as _json
    params_by_id: dict[int, dict | None] = {}
    for sig in group_signals:
        try:
            params_by_id[sig.id] = _json.loads(sig.params)
        except (TypeError, ValueError):
            params_by_id[sig.id] = None

    gscores = (
        s.query(GroupScore)
        .filter(GroupScore.date == target_date)
        .all()
    )

    existing_gsvs: dict[tuple, GroupSignalValue] = {
        (gsv.signal_id, gsv.group_type, gsv.group_id): gsv
        for gsv in s.query(GroupSignalValue).filter(GroupSignalValue.date == target_date).all()
    }

    scores = _evaluate_group_signal_scores(
        group_signals=group_signals, params_by_id=params_by_id, gscores=gscores)

    written = 0
    for (sig_id, group_type, group_id), score in scores.items():
        key = (sig_id, group_type, group_id)
        gsv = existing_gsvs.get(key)
        if gsv is None:
            gsv = GroupSignalValue(
                signal_id=sig_id,
                group_type=group_type,
                group_id=group_id,
                date=target_date,
            )
            s.add(gsv)
            existing_gsvs[key] = gsv
        gsv.score = score
        written += 1

    s.commit()
    logger.info("signal_service: %d group_signal_value escritos para %s", written, target_date)
    return written


def _evaluate_group_signal_scores(*, group_signals, params_by_id,
                                  gscores) -> dict[tuple, float]:
    """{(signal_id, group_type, group_id): score} de una fecha — LÓGICA
    PURA compartida por el camino por-fecha y el modo rango. gscores:
    iterable de objetos con group_type/group_id/regime_score_*."""
    results: dict[tuple, float] = {}
    for gscore in gscores:
        for sig in group_signals:
            if sig.group_type and sig.group_type != gscore.group_type:
                continue
            value = _get_group_indicator_value(gscore, sig.indicator_key) if sig.indicator_key else None
            score = signal_engine.evaluate(sig.formula_type, sig.params, value,
                                           params=params_by_id.get(sig.id))
            if score is None:
                continue
            results[(sig.id, gscore.group_type, gscore.group_id)] = score
    return results


def run_daily(target_date: date_type | None = None) -> dict:
    if target_date is None:
        from app.services.group_score_service import get_default_target_date
        target_date = get_default_target_date()

    asset_written = compute_signal_values(target_date)
    group_written = compute_group_signal_values(target_date)

    return {"date": str(target_date), "signal_values": asset_written, "group_signal_values": group_written}


# ── CRUD ──────────────────────────────────────────────────────────────────────

def get_all_signals() -> list:
    s = get_session()
    return s.query(SignalDefinition).order_by(SignalDefinition.id).all()


def get_visible_signals(user_id: int | None, is_admin: bool) -> list:
    """Señales visibles para el usuario: públicas + propias (admin: todas).
    Para pantallas y dropdowns — el pipeline de cálculo usa get_all_signals."""
    from app.services.visibility import visible_filter
    s = get_session()
    return (s.query(SignalDefinition)
            .filter(visible_filter(SignalDefinition, user_id, is_admin))
            .order_by(SignalDefinition.id).all())


def _signal_dependents(s, sig: SignalDefinition) -> list[tuple]:
    """[(descripción, owner_id, is_public)] de las estrategias que referencian a
    sig (componentes u operandos señal del filtro de elegibilidad)."""
    from app.models import Strategy
    from app.services import strategy_filter

    deps: list[tuple] = []

    for strat in s.query(Strategy).all():
        in_components = any(c.signal_id == sig.id for c in strat.components)
        in_filter = False
        tree = strategy_filter.parse_tree(strat.filter_conditions)
        if tree is not None:
            in_filter = any(t == "signal" and k == sig.key
                            for t, k, _res in strategy_filter.collect_operands(tree))
        if in_components or in_filter:
            deps.append((f"estrategia '{strat.name}'",
                         strat.owner_id, strat.is_public))

    return deps


def affected_by_signal_change(signal_id: int) -> list[str]:
    """Descripciones de lo que queda desactualizado al EDITAR una señal: las
    estrategias que la usan (componentes o filtro). Para el aviso de
    "Recalcular completo" — no incluye la propia señal, el llamador la
    antepone."""
    s = get_session()
    sig = s.query(SignalDefinition).filter(SignalDefinition.id == signal_id).first()
    if sig is None:
        return []
    return [desc for desc, _owner, _public in _signal_dependents(s, sig)]


def signal_dependents_of_others(s, sig: SignalDefinition,
                                owner_id: int | None) -> list[str]:
    """Dependientes de sig que NO son privados del mismo dueño (públicos o
    de otro usuario). Si esta lista no está vacía, despublicar sig los
    dejaría apuntando a algo que sus dueños ya no ven."""
    return [desc for desc, dep_owner, dep_public in _signal_dependents(s, sig)
            if dep_public or dep_owner != owner_id]


_GROUP_TYPE_COLUMNS = {
    "sector": "sector_id", "market": "market_id", "industry": "industry_id",
    "country": "country_id", "instrument_type": "instrument_type_id",
}


def signals_and_strategies_affected_by_new_assets(asset_ids) -> list[str]:
    """Al AGREGAR activos (p.ej. sintéticos de conversión de moneda), sus
    señales/estrategias PROPIAS entran solas en la próxima corrida, pero los
    activos nuevos también pasan a integrar los AGREGADOS de sus grupos
    (sector/mercado/país/...): eso desactualiza en la historia las señales de
    grupo de esos tipos y las estrategias que las usan (su score/ranking
    histórico se calculó sin estos activos). Devuelve descripciones para el
    aviso de 'Recalcular completo'. Lista vacía = nada que recalcular (típico
    si no hay señales de grupo). No es transversal por activo: solo mira los
    grupos que los nuevos activos tocan."""
    if not asset_ids:
        return []
    s = get_session()

    types: set[str] = set()
    for a in s.query(Asset).filter(Asset.id.in_(list(asset_ids))).all():
        for gt, col in _GROUP_TYPE_COLUMNS.items():
            if getattr(a, col) is not None:
                types.add(gt)
    if not types:
        return []

    group_sigs = s.query(SignalDefinition).filter(
        SignalDefinition.source == "group",
        SignalDefinition.group_type.in_(types)).all()
    if not group_sigs:
        return []

    # Sets: una estrategia que usa dos señales de grupo afectadas se lista una
    # sola vez; los nombres de señal son únicos por key
    signal_descs = {f"señal de grupo «{sig.key}»" for sig in group_sigs}
    strat_descs: set[str] = set()
    for sig in group_sigs:
        for desc, _o, _p in _signal_dependents(s, sig):
            strat_descs.add(desc)
    return sorted(signal_descs) + sorted(strat_descs)


def save_signal(
    key: str,
    name: str,
    source: str,
    formula_type: str,
    params_json: str,
    *,
    description: str | None = None,
    group_type: str | None = None,
    indicator_key: str | None = None,
    signal_id: int | None = None,
    is_public: bool | None = None,
    acting_user_id: int | None = None,
    acting_is_admin: bool = True,
) -> SignalDefinition:
    """is_public None = conservar el valor actual (o privada si es nueva).
    acting_* identifican a quién guarda: en alta queda como dueño; en
    edición se valida el permiso (default admin para scripts/tests)."""
    import json as _json
    from app.services.visibility import can_edit

    if formula_type == "composite":
        raise ValueError(
            "La fórmula compuesta se removió: combiná señales en la estrategia, "
            "con componentes ponderados.")

    params = _json.loads(params_json)
    shape_error = signal_engine.validate_params(
        formula_type, params if isinstance(params, dict) else {})
    if shape_error:
        raise ValueError(shape_error)

    s = get_session()

    # Validar referencias (mismos chequeos que el import): una señal que
    # apunta a un indicador inexistente o inválido para su fuente guarda
    # bien y después nunca puntúa, silenciosamente
    if source == "group":
        if not group_type:
            raise ValueError("Una señal de grupo requiere tipo de grupo.")
        if indicator_key not in _VALID_GROUP_INDICATOR_KEYS:
            raise ValueError(
                f"indicator_key '{indicator_key}' no es un campo de "
                f"group_scores (válidos: "
                f"{sorted(_VALID_GROUP_INDICATOR_KEYS)}).")
    elif indicator_key and indicator_key not in _VIRTUAL_CODES:
        from app.models.indicator_definition import IndicatorDefinition
        known = s.query(IndicatorDefinition.id).filter(
            IndicatorDefinition.code == indicator_key).first()
        if known is None:
            raise ValueError(f"Indicador desconocido: '{indicator_key}'.")
    if signal_id:
        sig = s.query(SignalDefinition).filter(SignalDefinition.id == signal_id).first()
        if sig is None:
            raise ValueError(f"Señal id={signal_id} no encontrada.")
        if not can_edit(sig.owner_id, acting_user_id, acting_is_admin):
            raise ValueError("Solo el dueño o un administrador pueden "
                             "editar esta señal.")
        new_public = sig.is_public if is_public is None else bool(is_public)
        if sig.is_public and not new_public:
            deps = signal_dependents_of_others(s, sig, sig.owner_id)
            if deps:
                raise ValueError(
                    "No se puede despublicar: la referencian "
                    + ", ".join(sorted(set(deps))) + ".")
    else:
        existing = s.query(SignalDefinition).filter(
            db_compat.ci_equals(SignalDefinition.key, key)).first()
        if existing:
            raise ValueError(f"Ya existe una señal con key '{key}'.")
        sig = SignalDefinition()
        sig.owner_id = acting_user_id
        s.add(sig)
        new_public = bool(is_public)

    sig.key           = key
    sig.name          = name
    sig.description   = description
    sig.source        = source
    sig.group_type    = group_type or None
    sig.indicator_key = indicator_key or None
    sig.formula_type  = formula_type
    sig.params        = params_json
    sig.is_public     = new_public
    try:
        s.commit()
    except Exception:
        s.rollback()
        raise
    # DESPUÉS del commit (orden ante crash: definición sin tabla es el lado
    # benigno — cualquier ensure_* posterior la repara). El id es el nombre
    # de la tabla, inmutable: renombrar la key no toca el almacenamiento.
    signal_store.ensure_sig_table(sig.id)
    return sig


def delete_signal(signal_id: int, *, acting_user_id: int | None = None,
                  acting_is_admin: bool = True) -> None:
    from app.services.visibility import can_edit
    s = get_session()
    sig = s.query(SignalDefinition).filter(SignalDefinition.id == signal_id).first()
    if sig is None:
        raise ValueError(f"Señal id={signal_id} no encontrada.")
    if not can_edit(sig.owner_id, acting_user_id, acting_is_admin):
        raise ValueError(f"Solo el dueño o un administrador pueden "
                         f"eliminar la señal '{sig.key}'.")
    # Chequeo ANTES de borrar: el FK de strategy_component igual lo
    # impediría, pero con un IntegrityError críptico que además deja la
    # sesión en estado rolled-back (envenena los requests siguientes)
    deps = [desc for desc, _o, _p in _signal_dependents(s, sig)]
    if deps:
        raise ValueError(
            f"No se puede eliminar '{sig.key}': la usan "
            + ", ".join(sorted(set(deps)))
            + ". Quitala de ahí primero.")
    from app.models import SignalEvalLog
    sig_id = sig.id
    try:
        s.query(SignalEvalLog).filter(
            SignalEvalLog.scope_kind == "signal",
            SignalEvalLog.ref_id == sig.id).delete()
        s.delete(sig)
        s.commit()
    except Exception:
        s.rollback()
        raise
    # DROP después del commit: si crashea en el medio queda una tabla
    # huérfana inofensiva (la barre reconcile_dynamic_tables), nunca una
    # definición sin tabla. DROP es instantáneo — antes el CASCADE de
    # signal_value borraba millones de filas reteniendo locks.
    signal_store.drop_sig_table(sig_id)


# ── Export / Import Excel ──────────────────────────────────────────────────────

def export_signals_excel() -> bytes:
    import openpyxl
    from io import BytesIO
    from app.services.visibility import publica_str

    signals = get_all_signals()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Señales"
    ws.append(["key", "name", "description", "source", "group_type",
                "indicator_key", "formula_type", "params", "publica"])
    for sig in signals:
        ws.append([
            sig.key, sig.name, sig.description or "",
            sig.source, sig.group_type or "", sig.indicator_key or "",
            sig.formula_type, sig.params, publica_str(sig.is_public),
        ])
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def import_signals_excel(file_bytes: bytes,
                         owner_id: int | None = None) -> list[dict]:
    """Importación todo-o-nada en dos pasadas: primero se valida el archivo
    completo sin tocar la base; solo si no hay errores se escribe todo.

    La columna `publica` (sí/no; ausente = PRIVADA) define la visibilidad de
    cada fila. owner_id = quien importa: queda como dueño de las señales
    NUEVAS (las existentes conservan su dueño)."""
    import openpyxl
    import json as _json
    from io import BytesIO
    from app.services.visibility import parse_publica

    wb = openpyxl.load_workbook(BytesIO(file_bytes))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(h).strip().lower() for h in rows[0]]

    _FORMULA_TYPES = ("discrete_map", "threshold", "range")
    _SOURCES       = ("asset", "group")

    # Catálogos para validar referencias (indicadores, señales existentes)
    s = get_session()
    from app.models.indicator_definition import IndicatorDefinition
    known_indicators = {
        d.code for d in s.query(IndicatorDefinition.code).all()
    } | set(_VIRTUAL_CODES)
    db_signals = {sig.key: sig for sig in s.query(SignalDefinition).all()}

    # ── Pasada 1: validación completa sin escribir ────────────────────────────
    parsed: list[dict] = []
    invalid = False
    for row in rows[1:]:
        data = dict(zip(headers, row))
        key = str(data.get("key") or "").strip()
        if not key:
            continue
        params_str    = str(data.get("params") or "{}")
        formula_type  = str(data.get("formula_type") or "range")
        source        = str(data.get("source") or "asset")
        indicator_key = str(data.get("indicator_key") or "").strip()
        group_type    = str(data.get("group_type") or "").strip()
        error = None
        params = None
        is_public = False   # solo se usa si parse_publica lanza; default privado
        try:
            is_public = parse_publica(data.get("publica"))
        except ValueError as exc:
            error = str(exc)
        try:
            params = _json.loads(params_str)
        except Exception as exc:
            error = error or f"params inválido: {exc}"
        if error is None and formula_type not in _FORMULA_TYPES:
            error = f"formula_type desconocido: '{formula_type}'"
        if error is None and source not in _SOURCES:
            error = f"source desconocido: '{source}'"
        if error is None:
            # Forma de params según la fórmula: un params json-válido pero
            # con la forma equivocada no rompe nada — la señal nunca
            # puntuaría, silenciosamente. Mejor rechazar acá.
            error = signal_engine.validate_params(
                formula_type, params if isinstance(params, dict) else {})
        if error is None:
            if source == "group":
                if not group_type:
                    error = "source=group requiere group_type"
                elif indicator_key not in _VALID_GROUP_INDICATOR_KEYS:
                    error = (f"indicator_key '{indicator_key}' no es un campo "
                             f"válido de group_scores")
            elif indicator_key and indicator_key not in known_indicators:
                error = f"indicador desconocido: '{indicator_key}'"
        if error:
            invalid = True
        parsed.append({"key": key, "data": data, "params": params_str,
                       "formula_type": formula_type, "source": source,
                       "is_public": is_public, "error": error})

    # Despublicar vía import: mismo chequeo de dependientes que en el ABM
    for p in parsed:
        if p["error"] is None:
            existing = db_signals.get(p["key"])
            if existing is not None and existing.is_public and not p["is_public"]:
                deps = signal_dependents_of_others(s, existing, existing.owner_id)
                if deps:
                    p["error"] = ("no se puede despublicar: la referencian "
                                  + ", ".join(sorted(set(deps))))
                    invalid = True

    if invalid:
        return [
            {"key": p["key"],
             "status": "error" if p["error"] else "omitido",
             "detail": p["error"] or "el archivo contiene errores; no se importó nada"}
            for p in parsed
        ]

    # ── Pasada 2: escribir todo en una sola transacción ───────────────────────
    # Upsert por key: una señal repetida entre archivos (packs
    # autosuficientes que comparten señales) NUNCA se duplica — actualiza la
    # misma fila. El detail distingue creada/actualizada/sin cambios.
    s = get_session()
    results: list[dict] = []
    try:
        for p in parsed:
            data = p["data"]
            key  = p["key"]
            # ci_equals: en MySQL el match por key ya era case-insensitive
            # (collation) — sin esto, en PG un re-import con otro caso
            # duplicaría la señal en vez de actualizarla
            sig = s.query(SignalDefinition).filter(
                db_compat.ci_equals(SignalDefinition.key, key)).first()
            new_vals = dict(
                key=key,
                name=str(data.get("name") or key),
                source=p["source"],
                formula_type=p["formula_type"],
                params=p["params"],
                description=str(data.get("description") or "") or None,
                group_type=str(data.get("group_type") or "") or None,
                indicator_key=str(data.get("indicator_key") or "") or None,
                is_public=p["is_public"],
            )
            if sig is None:
                sig = SignalDefinition()
                sig.owner_id = owner_id
                s.add(sig)
                outcome = "creada"
            elif all(getattr(sig, f) == v for f, v in new_vals.items()):
                outcome = "ya existía, sin cambios"
            else:
                outcome = "actualizada"
            for f, v in new_vals.items():
                setattr(sig, f, v)
            s.flush()
            results.append({"key": key, "status": "ok",
                            "detail": f"{outcome} (id={sig.id})",
                            "_sig_id": sig.id})
        s.commit()
    except Exception as exc:
        s.rollback()
        failed_key = parsed[len(results)]["key"] if len(results) < len(parsed) else "?"
        return [
            {"key": p["key"],
             "status": "error" if p["key"] == failed_key else "revertido",
             "detail": str(exc) if p["key"] == failed_key
                       else "revertido por error en otra fila"}
            for p in parsed
        ]

    # Tablas sig_{id} después del commit (mismo orden ante crash que
    # save_signal); ensure es idempotente para las ya existentes
    for r in results:
        signal_store.ensure_sig_table(r.pop("_sig_id"))
    return results


# ── Backfill histórico (Centro de Datos) ──────────────────────────────────────

# Umbral para cambiar del loop por-fecha al modo rango (signal_backfill_range)
_RANGE_MODE_MIN_DATES = 30


def _dates_to_compute(trading_dates: list, computed_dates: set,
                      force: bool) -> list:
    """Qué fechas correr. force=False (delta): las que no tienen ningún
    signal_value, más SIEMPRE la última (sus precios/indicadores son
    preliminares — mismo criterio que el delta de indicadores). force=True:
    todas."""
    if force:
        return list(trading_dates)
    if not trading_dates:
        return []
    last = trading_dates[-1]
    return [d for d in trading_dates if d not in computed_dates or d == last]


def _scope_signal_ids(s, scope: str | None):
    """Resuelve el alcance del backfill.

    scope: None (todo) | "strategy:<id>" | "signal:<key>".
    Devuelve (only_signal_ids | None, strategy_id | None, scope_kind).
    Para una estrategia incluye: señales de sus componentes + señales usadas
    como operando en su filtro de elegibilidad."""
    if not scope:
        return None, None, None
    kind, _, val = scope.partition(":")
    signals = s.query(SignalDefinition).all()
    by_key = {sg.key: sg for sg in signals}

    if kind == "strategy":
        from app.models import Strategy
        from app.services import strategy_filter
        strategy_id = int(val)
        strat = s.query(Strategy).filter(Strategy.id == strategy_id).first()
        if strat is None:
            raise ValueError(f"Estrategia id={strategy_id} no encontrada.")
        seed = {c.signal_id for c in strat.components}
        tree = strategy_filter.parse_tree(strat.filter_conditions)
        if tree is not None:
            for t, key, _res in strategy_filter.collect_operands(tree):
                if t == "signal" and key in by_key:
                    seed.add(by_key[key].id)
        return seed, strategy_id, "strategy"

    if kind == "signal":
        sig = by_key.get(val)
        if sig is None:
            raise ValueError(f"Señal '{val}' no encontrada.")
        return {sig.id}, None, "signal"

    raise ValueError(f"Alcance desconocido: {scope!r}")


def _signal_history_run(progress_cb=None, days: int | None = None,
                        force: bool = False, scope: str | None = None,
                        with_signals: bool = True) -> dict:
    """Corre el pipeline (scores de grupo → señales → estrategias) para cada
    fecha con precios dentro del horizonte. Delta (force=False): solo fechas
    sin cálculo previo + la última. Rebuild (force=True): todas.

    days None o 0 = SIN horizonte (toda la historia de precios). En delta
    es barato (las queries de fechas usan el índice por date y el trabajo
    real es solo sobre los huecos); en rebuild recalcula la historia entera.

    scope acota a una estrategia (señales necesarias + sus resultados) o a
    una señal suelta (sin tocar resultados de estrategias).

    with_signals=False (solo con scope de estrategia — decisión del USUARIO
    cuando no cambiaron señales ni indicadores): las señales no se
    re-evalúan ni se reescriben — sus scores se leen de signal_value — y
    solo se reconstruye strategy_result. Si cambió un indicador o una
    señal, corresponde with_signals=True (el pipeline completo)."""
    from datetime import timedelta

    from app.services import group_score_service, strategy_service

    s = get_session()
    only_ids, strategy_id, scope_kind = _scope_signal_ids(s, scope)

    last = s.query(sa.func.max(Price.date)).scalar()
    if last is None:
        return {"total": 0, "success": 0, "errors": [], "unit": "fechas"}
    horizon = (last - timedelta(days=int(days))) if days else None

    def _within(q, col):
        return q.filter(col >= horizon) if horizon is not None else q

    trading_dates = sorted({
        d for (d,) in _within(s.query(Price.date).distinct(), Price.date).all()
    })

    # "Ya calculado" según el alcance: los resultados de ESA estrategia, los
    # scores de ESA señal, o cualquier señal del día (alcance total).
    # En force (rebuild) NO se consulta: _dates_to_compute recorre todas las
    # fechas sin mirar `computed` — el DISTINCT sobre las tablas grandes
    # (18s por estrategia; minutos sobre signal_value entero en el alcance
    # total) era arranque 100% tirado.
    def _distinct_dates(table_or_col, extra_filter=None) -> set:
        """DISTINCT date de una tabla dinámica (Table core) dentro del
        horizonte — sobre el prefijo de la PK (date, asset_id) es un loose
        index scan barato."""
        q = sa.select(table_or_col.c.date).distinct()
        if extra_filter is not None:
            q = q.where(extra_filter)
        if horizon is not None:
            q = q.where(table_or_col.c.date >= horizon)
        return {d for (d,) in s.execute(q)}

    if scope_kind == "strategy":
        eval_kind, eval_ref = "strategy", strategy_id
        t = signal_store.ensure_strat_table(strategy_id, bind=s.connection())
        computed = set() if force else _distinct_dates(t)
    elif scope_kind == "signal":
        # La señal pedida (no sus dependencias) define el "ya calculado"
        target_sig = s.query(SignalDefinition).filter(
            db_compat.ci_equals(SignalDefinition.key,
                                scope.partition(":")[2])).first()
        eval_kind, eval_ref = "signal", target_sig.id
        if force:
            computed = set()
        elif target_sig.source == "group":
            computed = {
                d for (d,) in _within(
                    s.query(GroupSignalValue.date).distinct().filter(
                        GroupSignalValue.signal_id == target_sig.id),
                    GroupSignalValue.date)
            }
        else:
            computed = _distinct_dates(
                signal_store.ensure_sig_table(target_sig.id,
                                              bind=s.connection()))
    else:
        eval_kind, eval_ref = "all", 0
        computed = set()
        if not force:
            # Unión de fechas de todas las tablas sig_{id} (pocas señales,
            # una query barata por tabla)
            for (sid,) in s.query(SignalDefinition.id).filter(
                    SignalDefinition.source == "asset"):
                computed |= _distinct_dates(
                    signal_store.ensure_sig_table(sid, bind=s.connection()))

    # Fechas ya evaluadas que produjeron 0 filas (nadie pasó el filtro, señal
    # sin datos ese día): sin este registro parecen huecos y el delta las
    # reprocesa entero en CADA corrida (1927→1993 con solo ^GSPC, p. ej.)
    from app.models import SignalEvalLog
    logged = {
        d for (d,) in _within(
            s.query(SignalEvalLog.date).filter(
                SignalEvalLog.scope_kind == eval_kind,
                SignalEvalLog.ref_id == eval_ref),
            SignalEvalLog.date)
    }
    computed |= logged

    dates = _dates_to_compute(trading_dates, computed, force)

    # Modo rango: con muchas fechas, el loop por-fecha repite queries
    # constantes/incrementales 25.000 veces — el barrido cronológico hace lo
    # mismo con una carga por chunk (ver signal_backfill_range). El camino
    # por-fecha queda para el uso diario (última fecha, pocos huecos).
    strategy_only = (scope_kind == "strategy" and not with_signals)

    if len(dates) >= _RANGE_MODE_MIN_DATES:
        from app.services import signal_backfill_range
        return signal_backfill_range.run_range(
            dates,
            only_ids=only_ids, strategy_id=strategy_id,
            scope_kind=scope_kind, latest_price_date=last,
            eval_kind=eval_kind, eval_ref=eval_ref, logged=logged,
            progress_cb=progress_cb, force=force,
            full_wipe=(force and horizon is None and scope_kind is None),
            whole_history=(force and horizon is None),
            strategy_only=strategy_only)

    total, ok, errors = len(dates), 0, []
    for i, d in enumerate(dates, start=1):
        if progress_cb:
            progress_cb(i, total, str(d))
        try:
            if not strategy_only:
                group_score_service.run_daily(d)
                compute_signal_values(d, only_signal_ids=only_ids,
                                      latest_price_date=last)
                compute_group_signal_values(d, only_signal_ids=only_ids)
            # strategy_only: compute_strategy_results lee signal_value/
            # group_signal_value de la BD — exactamente la semántica pedida
            if scope_kind == "strategy":
                strategy_service.compute_strategy_results(strategy_id, d)
            elif scope_kind is None:
                strategy_service.compute_all_strategies(d)
            # scope señal: no toca resultados de estrategias
            if d not in logged:
                s.add(SignalEvalLog(scope_kind=eval_kind, ref_id=eval_ref, date=d))
                s.commit()
                logged.add(d)
            ok += 1
        except Exception as exc:
            logger.exception("signal_service: backfill falló para %s", d)
            errors.append({"date": str(d), "error": f"{d}: {exc}"})
            # Sin rollback, un fallo a mitad de un flush deja la sesión
            # compartida en estado rolled-back y TODAS las fechas
            # siguientes fallarían con "issue Session.rollback()"
            s.rollback()
    return {"total": total, "success": ok, "errors": errors, "unit": "fechas"}


def update_signal_history(progress_cb=None, days: int | None = None,
                          scope: str | None = None,
                          with_signals: bool = True) -> dict:
    """Delta: llena huecos (fechas con precios pero sin cálculo) y recalcula
    la última fecha. Default SIN horizonte (toda la historia): un hueco
    viejo nunca queda invisible — el trabajo real es solo sobre los huecos,
    así que el scheduler nocturno lo usa igual. days acota si se pide."""
    return _signal_history_run(progress_cb, days=days, force=False,
                               scope=scope, with_signals=with_signals)


def rebuild_signal_history(progress_cb=None, days: int | None = None,
                           scope: str | None = None,
                           with_signals: bool = True) -> dict:
    """Rebuild: recalcula TODAS las fechas con precios del horizonte
    (reescribe lo existente — para cambios de definición o fixes).
    days None/0 = toda la historia: puede tardar MUCHO (pipeline completo
    por cada fecha)."""
    return _signal_history_run(progress_cb, days=days, force=True,
                               scope=scope, with_signals=with_signals)


def run_recalculate(target_date: date_type | None = None) -> dict:
    from app.services import group_score_service, strategy_service

    if target_date is None:
        target_date = group_score_service.get_default_target_date()

    group_score_service.run_daily(target_date)
    result = run_daily(target_date)

    strat_result = strategy_service.run_daily(target_date)
    result["strategy_results"] = strat_result.get("strategy_results", 0)
    return result
