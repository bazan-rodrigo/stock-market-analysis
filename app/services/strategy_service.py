"""
Servicio de estrategias.
Combina los scores de señal ponderados según la configuración de
strategy_component y persiste el score final + ranking en strat_res_{id}
(una tabla por estrategia, ver app.models.signal_store).
"""
import logging
from datetime import date as date_type

import sqlalchemy as sa

from app.database import get_session
from app.models import (
    Asset,
    Strategy,
    StrategyComponent,
)
from app.models import signal_store
from app.services import db_compat, strategy_filter

logger = logging.getLogger(__name__)


def _compute_asset_score(
    components: list[StrategyComponent],
    asset_id: int,
    signal_scores: dict[tuple, float],
) -> float | None:
    """
    Calcula el score ponderado de una estrategia para un activo.

    signal_scores: {(signal_id, asset_id): score}

    El peso admite SIGNO: negativo = la señal aporta al revés (puntúa alto
    donde la señal puntúa bajo). Sirve para "momentum alto PERO volatilidad
    baja" sin duplicar la señal invertida en el catálogo.

    El divisor es Σ|peso|, NO Σpeso. La diferencia no es cosmética: con Σpeso
    un único componente de peso −1 daba −s/−1 = +s, o sea el peso negativo se
    cancelaba contra su propio denominador y NO invertía nada; y con pesos
    mixtos el divisor tendía a 0 y el score se disparaba fuera de −100..100
    (con +1 y −1 daba divisor 0 → None). Con Σ|peso| el resultado sigue siendo
    una combinación convexa: como cada score de señal está clampeado a
    −100..100 (signal_engine), el score de estrategia queda en ese rango, que
    es lo que da sentido a los umbrales ABSOLUTOS del simulador de trades
    (entries `score >= th`, salidas `absolute`/`delta_entry`/`trailing_score`).

    Retrocompatible: con todos los pesos positivos abs() es la identidad y el
    resultado no cambia. Solo cambia para pesos ≤ 0, que hasta ahora daban un
    resultado equivocado en silencio.
    """
    total_weight = 0.0
    weighted_sum = 0.0

    for comp in components:
        score = signal_scores.get((comp.signal_id, asset_id))
        if score is None:
            continue

        # `is None` y no `or`: con `or`, un peso 0 se convertía en 1.0 — un
        # componente que el usuario quiso anular pesaba como cualquier otro.
        weight = 1.0 if comp.weight is None else comp.weight
        weighted_sum += score * weight
        total_weight += abs(weight)

    if total_weight == 0:
        return None
    return round(weighted_sum / total_weight, 4)


def parse_component_weight(value) -> float:
    """Peso de un componente → float con SIGNO. Fuente única: la usan el ABM y
    el import, para que la planilla y la pantalla acepten exactamente lo mismo.

    Vacío/None = 1.0 (el default histórico). Negativo = la señal aporta al
    revés. El 0 se RECHAZA: no aporta al score y tampoco al divisor, así que es
    un componente que no hace nada — casi siempre es un error de tipeo. Antes
    se colaba y encima `or 1.0` lo convertía en 1.0, que es lo contrario de lo
    que quiso quien lo escribió.
    """
    if value is None or (isinstance(value, str) and not value.strip()):
        return 1.0
    try:
        w = float(value)
    except (TypeError, ValueError):
        raise ValueError(f"peso inválido: {value!r} (tiene que ser un número)")
    if w != w or w in (float("inf"), float("-inf")):      # NaN / ±inf
        raise ValueError(f"peso inválido: {value!r}")
    if w == 0:
        raise ValueError(
            "el peso no puede ser 0 (un componente con peso 0 no aporta nada): "
            "sacá el componente, o usá un peso NEGATIVO si querés que la señal "
            "aporte al revés")
    return w


def percent_ranks(values: list[float]) -> list[float]:
    """Percentil 0..100 de cada valor dentro de la lista (100 = mejor),
    alineado al orden de entrada — LÓGICA PURA compartida por el camino
    por-fecha y el modo rango (misma fila de strategy_result.pct desde
    ambos, ver test_signal_range_parity).

    Semántica de SQL PERCENT_RANK(): (rank − 1) / (n − 1) × 100, con
    RANK() para empates (comparten el rango mínimo). n=1 → 0.0 (igual que
    SQL). Se persiste porque derivarlo al leer es carísimo (la serie de un
    activo necesita la cross-section completa de cada fecha), mientras que
    acá la cross-section ya está en memoria."""
    n = len(values)
    if n == 0:
        return []
    if n == 1:
        return [0.0]
    order = sorted(range(n), key=lambda i: values[i])
    ranks = [0] * n
    i = 0
    while i < n:
        j = i
        while j + 1 < n and values[order[j + 1]] == values[order[i]]:
            j += 1
        for k in range(i, j + 1):
            ranks[order[k]] = i + 1  # RANK(): empates comparten el mínimo
        i = j + 1
    return [(r - 1) / (n - 1) * 100 for r in ranks]


def rank_strategy_assets(*, components, asset_groups, signal_scores,
                         filter_tree, operand_values,
                         ) -> list[tuple[int, float]]:
    """[(asset_id, score)] ordenado por score desc (el primero es el mejor) —
    LÓGICA PURA compartida por el camino por-fecha y el modo rango: filtro
    de elegibilidad + score ponderado + orden.

    asset_groups: {asset_id: {atributo: id}} — enumera los activos candidatos
    y alimenta las condiciones de atributo del filtro de elegibilidad."""
    asset_ids = list(asset_groups.keys())
    if filter_tree is not None and asset_ids:
        passing = strategy_filter.evaluate_tree_bulk(
            filter_tree, asset_ids, operand_values, asset_groups)
        # Preserva el orden original (desempate estable del sort por score)
        asset_ids = [aid for aid in asset_ids if aid in passing]

    scored: list[tuple[int, float]] = []
    for asset_id in asset_ids:
        score = _compute_asset_score(components, asset_id, signal_scores)
        if score is not None:
            scored.append((asset_id, score))

    # Orden: mayor score primero
    scored.sort(key=lambda x: x[1], reverse=True)
    return scored


def compute_strategy_results(strategy_id: int, target_date: date_type) -> int:
    """
    Calcula StrategyResult para todos los activos para strategy_id y target_date.
    Devuelve cantidad de resultados escritos.
    """
    s = get_session()

    strategy = s.query(Strategy).filter(Strategy.id == strategy_id).first()
    if strategy is None:
        logger.warning("strategy_service: strategy_id=%d no encontrada", strategy_id)
        return 0

    components = strategy.components
    if not components:
        return 0

    signal_ids = list({c.signal_id for c in components})

    # Cargar los scores de señal relevantes del día
    signal_scores: dict[tuple, float] = {}
    if signal_store.use_wide_signal_tables():
        signal_store.ensure_wide_signal_tables(bind=s.connection())
        for _dt, aid, sid, score in signal_store.load_wide_signal_scores(
                s, signal_ids, target_date, target_date):
            signal_scores[(sid, aid)] = score
    else:
        for sig_id in signal_ids:  # una tabla por señal
            t = signal_store.ensure_sig_table(sig_id, bind=s.connection())
            for aid, score in s.execute(
                    sa.select(t.c.asset_id, t.c.score)
                    .where(t.c.date == target_date)):
                signal_scores[(sig_id, aid)] = score

    # Solo cargar grupos de activos que aparecen en los datos de señales del día
    # (los atributos alimentan el filtro de elegibilidad; el score sale de
    # signal_scores, no de agregados de grupo)
    asset_ids_with_data = list({asset_id for _, asset_id in signal_scores})
    if asset_ids_with_data:
        q = strategy_filter.asset_attributes_query(s).filter(
            Asset.id.in_(asset_ids_with_data))
        asset_groups: dict[int, dict] = {
            a.id: strategy_filter.attributes_from_asset_row(a) for a in q.all()
        }
    else:
        asset_groups = {}

    filter_tree = strategy_filter.parse_tree(strategy.filter_conditions)
    operand_values = (
        strategy_filter.load_operand_values(s, filter_tree, target_date)
        if filter_tree is not None and asset_groups else {}
    )
    scored = rank_strategy_assets(
        components=components, asset_groups=asset_groups,
        signal_scores=signal_scores,
        filter_tree=filter_tree, operand_values=operand_values)

    if signal_store.use_wide_signal_tables():
        # Camino ancho: dos columnas (score, pct) de esta estrategia en
        # strategy_results_wide. NULL de la fecha + UPSERT deja la fila
        # consistente (borra filas obsoletas) sin tocar otras estrategias.
        conn = s.connection()
        signal_store.ensure_wide_signal_tables(bind=conn)
        signal_store.ensure_strat_columns(strategy_id, bind=conn)
        d_str = str(target_date)
        cols = signal_store.strat_columns([strategy_id])
        pcts = percent_ranks([score for _, score in scored])
        signal_store.wide_null_columns(
            s, signal_store.STRAT_WIDE_TABLE, cols, [d_str])
        signal_store.wide_upsert(
            s, signal_store.STRAT_WIDE_TABLE, cols,
            signal_store.strat_wide_rows(
                {strategy_id: [(aid, d_str, score, pct)
                               for (aid, score), pct in zip(scored, pcts)]},
                [strategy_id]))
        s.commit()
        logger.info("strategy_service: %d resultados escritos (ancho) para "
                    "strategy_id=%d en %s", len(scored), strategy_id,
                    target_date)
        return len(scored)

    # Upsert del día en la tabla propia strat_res_{id}
    rt = signal_store.ensure_strat_table(strategy_id, bind=s.connection())
    existing: dict[int, tuple] = {
        aid: (score, pct) for aid, score, pct in s.execute(
            sa.select(rt.c.asset_id, rt.c.score, rt.c.pct)
            .where(rt.c.date == target_date))
    }

    # Eliminar resultados de un recálculo previo del mismo día cuyos activos
    # ya no obtienen score (filas obsoletas)
    scored_ids = {asset_id for asset_id, _ in scored}
    stale = [aid for aid in existing if aid not in scored_ids]
    if stale:
        s.execute(rt.delete().where(rt.c.date == target_date,
                                    rt.c.asset_id.in_(stale)))

    written = 0
    pcts = percent_ranks([score for _, score in scored])
    ins, upd = [], []
    for (asset_id, score), pct in zip(scored, pcts):
        prev = existing.get(asset_id)
        if prev is None:
            ins.append({"asset_id": asset_id, "date": target_date,
                        "score": score, "pct": pct})
        elif prev != (score, pct):
            upd.append({"aid": asset_id, "sc": score, "pc": pct})
        written += 1
    if ins:
        s.execute(rt.insert(), ins)
    if upd:
        s.execute(
            rt.update().where(rt.c.date == target_date,
                              rt.c.asset_id == sa.bindparam("aid"))
            .values(score=sa.bindparam("sc"), pct=sa.bindparam("pc")),
            upd)

    s.commit()
    logger.info(
        "strategy_service: %d resultados escritos para strategy_id=%d en %s",
        written, strategy_id, target_date,
    )
    return written


def compute_all_strategies(target_date: date_type) -> dict:
    """Calcula los resultados de todas las estrategias para target_date."""
    s = get_session()
    strategies = s.query(Strategy.id).all()
    total = 0
    for (sid,) in strategies:
        total += compute_strategy_results(sid, target_date)
    return {"date": str(target_date), "strategy_results": total}


def run_daily(target_date: date_type | None = None) -> dict:
    """Pipeline diario de estrategias."""
    if target_date is None:
        from app.services.group_score_service import get_default_target_date
        target_date = get_default_target_date()

    return compute_all_strategies(target_date)


# ── CRUD ──────────────────────────────────────────────────────────────────────

def get_all_strategies() -> list:
    s = get_session()
    return s.query(Strategy).order_by(Strategy.id).all()


def get_visible_strategies(user_id: int | None, is_admin: bool) -> list:
    """Estrategias visibles para el usuario: públicas + propias (admin:
    todas). Para pantallas y dropdowns — el pipeline usa get_all_strategies."""
    from app.services.visibility import visible_filter
    s = get_session()
    return (s.query(Strategy)
            .filter(visible_filter(Strategy, user_id, is_admin))
            .order_by(Strategy.id).all())


def _validate_signal_refs_visibility(s, *, owner_id, is_public,
                                     signal_keys: set[str]) -> None:
    """Estrategia pública solo referencia señales públicas; privada,
    públicas + del mismo dueño (componentes y operandos del filtro)."""
    from app.models import SignalDefinition
    from app.services.visibility import can_reference

    if not signal_keys:
        return
    for ref in s.query(SignalDefinition).filter(
            SignalDefinition.key.in_(signal_keys)).all():
        if not can_reference(owner_id, is_public, ref.owner_id, ref.is_public):
            if is_public:
                raise ValueError(
                    f"Una estrategia pública no puede usar la señal "
                    f"privada '{ref.key}' — publicala primero.")
            raise ValueError(
                f"No podés usar la señal privada '{ref.key}' de otro usuario.")


def _filter_signal_keys(filter_conditions: str | None) -> set[str]:
    """Keys de señal usadas como operando en el filtro de elegibilidad."""
    tree = strategy_filter.parse_tree(filter_conditions)
    if tree is None:
        return set()
    return {k for t, k, _res in strategy_filter.collect_operands(tree)
            if t == "signal" and k}


def get_strategy_by_id(strategy_id: int) -> Strategy | None:
    s = get_session()
    return s.query(Strategy).filter(Strategy.id == strategy_id).first()


def validate_filter_conditions(filter_conditions: str | None) -> list[str]:
    """Errores del árbol de condiciones contra los catálogos vigentes
    (indicadores, señales, valores discretos). Vacío si es válido o NULL."""
    import json
    from app.models import SignalDefinition
    from app.models.indicator_definition import IndicatorDefinition
    from app.services.indicator_catalog import CATEGORICAL_VALUES

    if not filter_conditions:
        return []
    try:
        tree = json.loads(filter_conditions)
    except (json.JSONDecodeError, TypeError):
        return ["filtro: JSON inválido"]
    if not tree:
        return []

    s = get_session()
    indicator_codes = {
        d.code: d.type
        for d in s.query(IndicatorDefinition.code, IndicatorDefinition.type).all()
    }
    signal_keys = {r.key for r in s.query(SignalDefinition.key).all()}
    return strategy_filter.validate_tree(
        tree,
        indicator_codes=indicator_codes,
        signal_keys=signal_keys,
        categorical_values=CATEGORICAL_VALUES,
    )


def save_strategy(
    name: str,
    components: list[dict],
    *,
    description: str | None = None,
    filter_conditions: str | None = None,
    strategy_id: int | None = None,
    is_public: bool | None = None,
    acting_user_id: int | None = None,
    acting_is_admin: bool = False,
) -> Strategy:
    """is_public None = conservar el valor actual (o privada si es nueva).
    acting_* identifican a quién guarda: en alta queda como dueño; en edición
    se valida el permiso con can_edit (dueño o admin — a diferencia de las
    señales, las estrategias las crea cualquiera).

    El default de acting_is_admin es False: FALLA CERRADO. Un caller que se
    olvide del flag queda como "no admin" y solo puede tocar lo propio, en vez
    de saltear el permiso por omisión. Los scripts y tests pasan True
    explícito. Lo fija tests/test_permisos_fallan_cerrado.py."""
    from datetime import datetime as _dt
    from app.models import SignalDefinition
    from app.services.visibility import can_edit

    filter_errors = validate_filter_conditions(filter_conditions)
    if filter_errors:
        raise ValueError("; ".join(filter_errors))

    s = get_session()
    if strategy_id:
        strat = s.query(Strategy).filter(Strategy.id == strategy_id).first()
        if strat is None:
            raise ValueError(f"Estrategia id={strategy_id} no encontrada.")
        if not can_edit(strat.owner_id, acting_user_id, acting_is_admin):
            raise ValueError("Solo el dueño o un administrador pueden "
                             "editar esta estrategia.")
        new_public = strat.is_public if is_public is None else bool(is_public)
        for comp in list(strat.components):
            s.delete(comp)
        s.flush()
    else:
        strat = Strategy()
        strat.owner_id   = acting_user_id
        strat.created_at = _dt.utcnow()
        s.add(strat)
        new_public = bool(is_public)

    _validate_signal_refs_visibility(
        s, owner_id=strat.owner_id, is_public=new_public,
        signal_keys={str(c.get("signal_key") or "").strip()
                     for c in components if c.get("signal_key")}
                    | _filter_signal_keys(filter_conditions))

    strat.name              = name
    strat.description       = description
    strat.filter_conditions = filter_conditions or None
    strat.is_public         = new_public
    strat.updated_at        = _dt.utcnow()
    s.flush()

    for comp_data in components:
        sig_key = str(comp_data.get("signal_key") or "").strip()
        if not sig_key:
            raise ValueError("Cada componente requiere signal_key.")
        sig = s.query(SignalDefinition).filter(
            db_compat.ci_equals(SignalDefinition.key, sig_key)).first()
        if sig is None:
            raise ValueError(f"Señal '{sig_key}' no encontrada.")
        comp = StrategyComponent(
            strategy_id=strat.id,
            signal_id=sig.id,
            weight=parse_component_weight(comp_data.get("weight")),
        )
        s.add(comp)

    try:
        s.commit()
    except Exception:
        s.rollback()
        raise
    # DESPUÉS del commit (ver signal_store: definición sin tabla/columna es el
    # lado benigno ante crash; el id inmutable es el nombre del almacenamiento)
    signal_store.ensure_strategy_storage(strat.id)
    return strat


def delete_strategy(strategy_id: int, *, acting_user_id: int | None = None,
                    acting_is_admin: bool = False) -> None:
    """Borra una estrategia (dueño o admin). Default cerrado — ver save_strategy."""
    from app.services.visibility import can_edit
    s = get_session()
    strat = s.query(Strategy).filter(Strategy.id == strategy_id).first()
    if strat is None:
        raise ValueError(f"Estrategia id={strategy_id} no encontrada.")
    if not can_edit(strat.owner_id, acting_user_id, acting_is_admin):
        raise ValueError(f"Solo el dueño o un administrador pueden "
                         f"eliminar la estrategia '{strat.name}'.")
    from app.models import SignalEvalLog
    try:
        s.query(SignalEvalLog).filter(
            SignalEvalLog.scope_kind == "strategy",
            SignalEvalLog.ref_id == strategy_id).delete()
        s.delete(strat)
        s.commit()
    except Exception:
        s.rollback()
        raise
    # DROP después del commit: un crash deja tabla/columna huérfana inofensiva
    # (reconcile la barre), nunca definición sin dato
    signal_store.drop_strategy_storage(strategy_id)


def get_strategy_results(strategy_id: int, target_date) -> list[dict]:
    s = get_session()
    rt = signal_store.read_strat_table(s, strategy_id)
    rows = s.execute(
        sa.select(rt.c.asset_id, Asset.ticker, Asset.name, rt.c.score)
        .join_from(rt, Asset, Asset.id == rt.c.asset_id)
        .where(rt.c.date == target_date)
        # NULLs al final en ambos motores (en PG un DESC puro los pondría
        # primero y encabezarían el ranking)
        .order_by(*db_compat.order_desc_nulls_last(rt.c.score))
    ).all()
    return [
        {"asset_id": aid, "ticker": ticker, "name": name, "score": score}
        for aid, ticker, name, score in rows
    ]


def get_strategy_results_with_breakdown(
    strategy_id: int,
    target_date,
    *,
    sector_id: int | None = None,
    market_id: int | None = None,
    limit: int | None = None,
) -> tuple[list[dict], list[dict], int]:
    """
    Devuelve (resultados, componentes, total) donde:
    - resultados: [{asset_id, ticker, name, sector_id, market_id, score,
                    delta_score, component_scores: {signal_key: score}}]
      ordenados por score desc (el orden ES el ranking).
    - componentes: [{signal_key, signal_name, weight}]
    - total: cuántos activos tiene el ranking completo, ignorando `limit`
      (para que la UI pueda decir cuántos quedaron afuera del tope).

    `limit` corta el ranking en el servidor: se queda con los primeros N por
    score. No es solo cosmética — recorta también las lecturas por señal y la
    de la fecha anterior, que van con un IN sobre los asset_id traídos.
    """
    from app.models import SignalDefinition

    s = get_session()
    strategy = s.query(Strategy).filter(Strategy.id == strategy_id).first()
    if strategy is None:
        return [], [], 0

    components = strategy.components
    sig_ids = [c.signal_id for c in components]

    sigs_by_id = {
        sig.id: sig
        for sig in s.query(SignalDefinition)
        .filter(SignalDefinition.id.in_(sig_ids))
        .all()
    }

    # Resultado base
    rt = signal_store.read_strat_table(s, strategy_id)
    q = (
        sa.select(rt.c.asset_id, rt.c.score, Asset.ticker, Asset.name,
                  Asset.sector_id, Asset.market_id)
        .join_from(rt, Asset, Asset.id == rt.c.asset_id)
        .where(rt.c.date == target_date)
    )
    if sector_id is not None:
        q = q.where(Asset.sector_id == sector_id)
    if market_id is not None:
        q = q.where(Asset.market_id == market_id)

    # El total se cuenta aparte solo cuando hay tope; sin tope las filas SON
    # el total y una query de más no compra nada.
    total = 0
    if limit is not None:
        total = s.execute(
            sa.select(sa.func.count()).select_from(q.subquery())
        ).scalar() or 0

    q = q.order_by(*db_compat.order_desc_nulls_last(rt.c.score))
    if limit is not None:
        q = q.limit(limit)
    rows = s.execute(q).all()
    if limit is None:
        total = len(rows)

    if not rows:
        return [], [], total

    asset_ids = [aid for aid, *_ in rows]

    sv_map: dict[tuple, float] = {}
    for sig_id in set(sig_ids):
        t = signal_store.read_sig_table(s, sig_id)
        for aid, score in s.execute(
                sa.select(t.c.asset_id, t.c.score)
                .where(t.c.date == target_date,
                       t.c.asset_id.in_(asset_ids))):
            sv_map[(sig_id, aid)] = score

    comp_meta = [
        {
            "signal_key":  sigs_by_id[c.signal_id].key  if c.signal_id in sigs_by_id else str(c.signal_id),
            "signal_name": sigs_by_id[c.signal_id].name if c.signal_id in sigs_by_id else "?",
            "weight":      c.weight,
        }
        for c in components
    ]

    # Fecha anterior con resultados para esta estrategia
    prev_date = s.execute(
        sa.select(sa.func.max(rt.c.date)).where(rt.c.date < target_date)
    ).scalar()

    prev_score_map: dict[int, float] = {}
    if prev_date:
        prev_score_map = {
            aid: score for aid, score in s.execute(
                sa.select(rt.c.asset_id, rt.c.score)
                .where(rt.c.date == prev_date,
                       rt.c.asset_id.in_(asset_ids)))
        }

    results = []
    for asset_id, r_score, ticker, name, s_id, m_id in rows:
        comp_scores: dict[str, float | None] = {}

        for comp in components:
            sig = sigs_by_id.get(comp.signal_id)
            key = sig.key if sig else str(comp.signal_id)
            comp_scores[key] = sv_map.get((comp.signal_id, asset_id))

        prev_sc   = prev_score_map.get(asset_id)
        delta_score = round(r_score - prev_sc, 4) if (prev_sc is not None and r_score is not None) else None

        results.append({
            "asset_id":    asset_id,
            "ticker":      ticker,
            "name":        name or "—",
            "sector_id":   s_id,
            "market_id":   m_id,
            "score":       r_score,
            "prev_score":  prev_sc,
            "delta_score": delta_score,
            "comp_scores": comp_scores,
        })

    return results, comp_meta, total


def get_filter_options(strategy_id: int, target_date) -> dict:
    """Devuelve opciones de sector y market para los activos con resultados."""
    from app.models import Sector, Market
    s = get_session()

    rt = signal_store.read_strat_table(s, strategy_id)
    asset_ids = [
        aid for (aid,) in s.execute(
            sa.select(rt.c.asset_id).where(rt.c.date == target_date))
    ]
    if not asset_ids:
        return {"sectors": [], "markets": []}

    sectors = (
        s.query(Asset.sector_id, Sector.name)
        .join(Sector, Sector.id == Asset.sector_id)
        .filter(Asset.id.in_(asset_ids), Asset.sector_id.isnot(None))
        .distinct()
        .order_by(Sector.name)
        .all()
    )
    markets = (
        s.query(Asset.market_id, Market.name)
        .join(Market, Market.id == Asset.market_id)
        .filter(Asset.id.in_(asset_ids), Asset.market_id.isnot(None))
        .distinct()
        .order_by(Market.name)
        .all()
    )
    return {
        "sectors": [{"label": n, "value": sid} for sid, n in sectors],
        "markets": [{"label": n, "value": mid} for mid, n in markets],
    }


def get_available_dates(strategy_id: int) -> list:
    """Devuelve las fechas con resultados para una estrategia, ordenadas desc."""
    s = get_session()
    rt = signal_store.read_strat_table(s, strategy_id)
    dates = s.execute(
        sa.select(rt.c.date).distinct().order_by(rt.c.date.desc())
    ).all()
    return [r[0] for r in dates]


def get_strategy_score_history(
    strategy_id: int,
    asset_ids: list[int],
    date_from=None,
    date_to=None,
) -> dict[int, list[tuple]]:
    """
    {asset_id: [(date, score), ...]} ordenado por fecha asc.
    """
    s = get_session()
    rt = signal_store.read_strat_table(s, strategy_id)
    q = sa.select(rt.c.asset_id, rt.c.date, rt.c.score).where(
        rt.c.asset_id.in_(asset_ids))
    if date_from:
        q = q.where(rt.c.date >= date_from)
    if date_to:
        q = q.where(rt.c.date <= date_to)
    q = q.order_by(rt.c.date)

    result: dict[int, list] = {aid: [] for aid in asset_ids}
    for asset_id, dt, score in s.execute(q):
        result[asset_id].append((dt, score))
    return result


def get_top_assets_for_strategy(strategy_id: int, target_date, limit: int = 20) -> list[dict]:
    """Top activos por score en target_date, para usar como sugerencia en el historial."""
    return get_strategy_results(strategy_id, target_date)[:limit]


# ── Export / Import Excel ──────────────────────────────────────────────────────

def export_strategies_excel() -> bytes:
    import openpyxl
    from io import BytesIO
    from app.models import SignalDefinition
    from app.services.visibility import publica_str

    strategies = get_all_strategies()
    wb = openpyxl.Workbook()

    from app.services.pack_service import COMPONENT_COLUMNS, STRATEGY_COLUMNS

    ws_s = wb.active
    ws_s.title = "Estrategias"
    ws_s.append(list(STRATEGY_COLUMNS))

    ws_c = wb.create_sheet("Componentes")
    ws_c.append(list(COMPONENT_COLUMNS))

    s = get_session()
    all_sig_ids = {comp.signal_id for strat in strategies for comp in strat.components}
    sigs_by_id = {
        sig.id: sig
        for sig in s.query(SignalDefinition).filter(SignalDefinition.id.in_(all_sig_ids)).all()
    } if all_sig_ids else {}

    for strat in strategies:
        ws_s.append([strat.name, strat.description or "",
                     strat.filter_conditions or "",
                     publica_str(strat.is_public)])
        for comp in strat.components:
            sig = sigs_by_id.get(comp.signal_id)
            ws_c.append([
                strat.name, sig.key if sig else "", comp.weight,
            ])

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def import_strategies_file(file_bytes: bytes, filename: str | None = None,
                           owner_id: int | None = None) -> list[dict]:
    """Punto de entrada de la pantalla: despacha por formato (pack JSON o
    planilla). Los dos caminos comparten validación y escritura."""
    from app.services import pack_service

    if pack_service.looks_like_json(file_bytes, filename):
        pack = pack_service.parse_pack(file_bytes)
        rows_s, rows_c = pack_service.strategy_rows_from_pack(pack)
    else:
        rows_s, rows_c = _strategy_rows_from_xlsx(file_bytes)
        if rows_s is None:
            return rows_c        # planilla equivocada: rows_c trae el error
    return import_strategy_rows(rows_s, rows_c, owner_id=owner_id)


def _strategy_rows_from_xlsx(file_bytes: bytes):
    """Planilla → (filas de estrategia, filas de componente).

    Si el archivo es la planilla equivocada devuelve (None, [resultado de
    error]) — el mismo formato de salida que el import, para que la pantalla
    lo muestre igual que cualquier otro rechazo.
    """
    import openpyxl
    from io import BytesIO

    wb = openpyxl.load_workbook(BytesIO(file_bytes))
    ws_s = wb.worksheets[0]
    rows_s = list(ws_s.iter_rows(values_only=True))
    if not rows_s:
        return [], []

    headers_s = [str(h).strip().lower() for h in rows_s[0]]
    # Guard: la planilla de SEÑALES comparte name/description/publica con la de
    # estrategias, así que subirla acá por error creaba "estrategias" con los
    # nombres de las señales y 0 componentes, en silencio. Las columnas
    # formula_type/indicator_key/params son propias de señales → rechazar claro.
    if {"formula_type", "indicator_key", "params"} & set(headers_s):
        return None, [{
            "name": "(archivo)", "status": "error",
            "detail": "esta planilla parece de SEÑALES (tiene columnas como "
                      "formula_type / indicator_key), no de estrategias. "
                      "Importala en la pantalla de Señales, o exportá la "
                      "plantilla de estrategias desde esta pantalla."}]

    filas_s = [dict(zip(headers_s, row)) for row in rows_s[1:]]

    filas_c: list[dict] = []
    if len(wb.worksheets) > 1:
        rows_c = list(wb.worksheets[1].iter_rows(values_only=True))
        headers_c = [str(h).strip().lower() for h in rows_c[0]] if rows_c else []
        filas_c = [dict(zip(headers_c, row)) for row in rows_c[1:]]
    return filas_s, filas_c


def import_strategies_excel(file_bytes: bytes,
                            owner_id: int | None = None) -> list[dict]:
    """Camino Excel (nombre histórico). Ver import_strategies_file."""
    rows_s, rows_c = _strategy_rows_from_xlsx(file_bytes)
    if rows_s is None:
        return rows_c
    return import_strategy_rows(rows_s, rows_c, owner_id=owner_id)


def import_strategy_rows(rows_s: list[dict], rows_c: list[dict],
                         owner_id: int | None = None) -> list[dict]:
    """Importación todo-o-nada, desde filas normalizadas (Excel o pack JSON).

    La columna `publica` (sí/no; ausente = PRIVADA) define la visibilidad.
    owner_id = quien importa: dueño de las estrategias NUEVAS (las
    existentes conservan el suyo)."""
    import json
    from datetime import datetime as _dt
    from app.models import SignalDefinition
    from app.services import pack_service
    from app.services.visibility import can_reference, parse_publica

    strategies: dict[str, dict] = {}
    for data in rows_s:
        name = str(data.get("name") or "").strip()
        if name:
            # Compatibilidad: Excel exportados antes de la migración 0061
            # traen "asset_filter" (JSON plano) en vez de "filter_conditions"
            filter_conditions = (
                data.get("filter_conditions")
                or strategy_filter.legacy_asset_filter_to_tree(
                    data.get("asset_filter"))
                or None
            )
            entry = {
                "description": data.get("description") or None,
                "filter_conditions": filter_conditions,
                "components": [],
                "is_public": False,   # default privado; parse_publica lo confirma
            }
            try:
                entry["is_public"] = parse_publica(data.get("publica"))
            except ValueError as exc:
                entry.setdefault("errors", []).append(str(exc))
            strategies[name] = entry

    for data in rows_c:
        sname = str(data.get("strategy_name") or "").strip()
        if sname not in strategies:
            continue
        try:
            comp = {
                "signal_key": str(data.get("signal_key") or "").strip(),
                "weight": parse_component_weight(data.get("weight")),
            }
            # El Alcance de grupo (scope) se removió: rechazar en vez de
            # descartarlo en silencio, que cambiaría el score de la estrategia.
            scope = str(data.get("scope") or "").strip()
            if scope:
                strategies[sname].setdefault("errors", []).append(
                    f"el alcance '{scope}' ya no se soporta: el Alcance de "
                    f"grupo se removió")
            strategies[sname]["components"].append(comp)
        except (TypeError, ValueError) as exc:
            strategies[sname].setdefault("errors", []).append(
                f"componente inválido: {exc}")

    db = get_session()

    # Atributos por NOMBRE → id de esta instalación. Los ids de catálogo son
    # distintos en cada base: un pack que los trajera fijos no sería portable,
    # así que el formato de intercambio usa nombres y se resuelven acá, una
    # sola vez para todo el archivo (ver pack_service).
    attr_index = pack_service.attribute_index(db)

    # ── Pasada 1: validación completa sin escribir ────────────────────────────
    all_keys = {
        c["signal_key"]
        for sdata in strategies.values()
        for c in sdata["components"] if c["signal_key"]
    }
    sigs_by_key = {
        r.key: r
        for r in db.query(SignalDefinition)
                   .filter(SignalDefinition.key.in_(all_keys)).all()
    } if all_keys else {}
    sig_ids_by_key = {k: r.id for k, r in sigs_by_key.items()}
    existing_by_name = {
        st.name: st for st in db.query(Strategy)
        .filter(Strategy.name.in_(strategies)).all()
    } if strategies else {}

    invalid = False
    for name, sdata in strategies.items():
        errors = sdata.setdefault("errors", [])
        existing = existing_by_name.get(name)
        sdata["owner_id"] = existing.owner_id if existing else owner_id
        # Una estrategia sin componentes no puntúa nada (el pipeline la saltea):
        # rechazarla en vez de crearla vacía en silencio. El caso típico es la
        # planilla sin la hoja "Componentes" (o con strategy_name que no matchea).
        if not sdata["components"]:
            errors.append(
                "estrategia sin componentes: agregá la hoja 'Componentes' "
                "(columnas strategy_name, signal_key, weight) con al menos una "
                "señal para esta estrategia")
        for comp in sdata["components"]:
            if not comp["signal_key"]:
                errors.append("componente sin signal_key")
            elif comp["signal_key"] not in sig_ids_by_key:
                errors.append(f"señal '{comp['signal_key']}' no encontrada")

        # Nombres de catálogo → ids ANTES de validar: validate_tree compara
        # los valores discretos contra el catálogo y un nombre sin resolver
        # llegaría hasta el evaluador, donde nunca matchearía ningún activo
        # (el filtro dejaría la estrategia vacía, sin error visible).
        tree = strategy_filter.parse_tree(sdata["filter_conditions"])
        if tree:
            tree, attr_errors = pack_service.resolve_attribute_values(
                tree, attr_index)
            errors.extend(attr_errors)
            if not attr_errors:
                sdata["filter_conditions"] = json.dumps(tree)
        errors.extend(validate_filter_conditions(sdata["filter_conditions"]))
        # Visibilidad de las señales referenciadas (componentes + filtro)
        ref_keys = ({c["signal_key"] for c in sdata["components"]
                     if c["signal_key"]}
                    | _filter_signal_keys(sdata["filter_conditions"]))
        for rk in sorted(ref_keys):
            ref = sigs_by_key.get(rk) or db.query(SignalDefinition).filter(
                db_compat.ci_equals(SignalDefinition.key, rk)).first()
            if ref is not None and not can_reference(
                    sdata["owner_id"], sdata["is_public"],
                    ref.owner_id, ref.is_public):
                errors.append(
                    f"estrategia {'pública' if sdata['is_public'] else 'privada'} "
                    f"usa la señal privada '{rk}' de otro dueño")
        if errors:
            invalid = True

    if invalid:
        return [
            {"name": name,
             "status": "error" if sdata["errors"] else "omitido",
             "detail": "; ".join(sdata["errors"])
                       or "el archivo contiene errores; no se importó nada"}
            for name, sdata in strategies.items()
        ]

    # ── Pasada 2: escribir todo en una sola transacción ───────────────────────
    results: list[dict] = []
    try:
        for name, sdata in strategies.items():
            # ci_equals: el upsert del import matchea por nombre sin
            # distinguir caso, como lo hacía la collation de MySQL
            existing = db.query(Strategy).filter(
                db_compat.ci_equals(Strategy.name, name)).first()
            if existing:
                strat = existing
                for comp in list(strat.components):
                    db.delete(comp)
                db.flush()
            else:
                strat = Strategy()
                strat.owner_id   = owner_id
                strat.created_at = _dt.utcnow()
                db.add(strat)

            strat.name              = name
            strat.description       = sdata["description"]
            strat.filter_conditions = sdata["filter_conditions"] or None
            strat.is_public         = sdata["is_public"]
            strat.updated_at        = _dt.utcnow()
            db.flush()

            for comp_data in sdata["components"]:
                db.add(StrategyComponent(
                    strategy_id=strat.id,
                    signal_id=sig_ids_by_key[comp_data["signal_key"]],
                    weight=comp_data["weight"],
                ))

            db.flush()
            results.append({"name": name, "status": "ok",
                            "detail": f"id={strat.id}", "_strat_id": strat.id})
        db.commit()
    except Exception as exc:
        db.rollback()
        names  = list(strategies)
        failed = names[len(results)] if len(results) < len(names) else "?"
        return [
            {"name": n,
             "status": "error" if n == failed else "revertido",
             "detail": str(exc) if n == failed
                       else "revertido por error en otra estrategia"}
            for n in names
        ]

    # Almacenamiento después del commit (mismo orden que save_strategy)
    for r in results:
        signal_store.ensure_strategy_storage(r.pop("_strat_id"))
    return results
